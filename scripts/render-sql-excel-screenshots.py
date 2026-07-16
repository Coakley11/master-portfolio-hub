#!/usr/bin/env python3
"""Render recruiter-facing SQL/Excel portfolio screenshots from workbook data."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

HUB = Path(__file__).resolve().parent.parent
SQL = HUB / "SQL Portfolio"
OUT = HUB / "Screenshots" / "SQL-Excel"
DEPLOY = HUB / "Portfolio Website" / "assets" / "screenshots" / "SQL-Excel"

W, H = 1600, 1000
BG = (15, 23, 42)
PANEL = (30, 41, 59)
CARD = (51, 65, 85)
TEXT = (226, 232, 240)
MUTED = (148, 163, 184)
ACCENT = (56, 189, 248)
GREEN = (52, 211, 153)
GOLD = (251, 191, 36)
RED = (248, 113, 113)


def font(size: int, bold: bool = False):
    candidates = [
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def sheet_rows(wb_path: Path, sheet: str, max_rows: int = 12, max_cols: int = 8):
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([("" if c is None else c) for c in row[:max_cols]])
    wb.close()
    return rows


def draw_table(draw, x, y, rows, col_w=150, row_h=28):
    if not rows:
        return y
    header = rows[0]
    for c, val in enumerate(header):
        draw.rectangle([x + c * col_w, y, x + (c + 1) * col_w - 2, y + row_h], fill=CARD)
        draw.text((x + c * col_w + 8, y + 6), str(val)[:18], fill=TEXT, font=font(14, True))
    y += row_h + 2
    for r in rows[1:]:
        for c, val in enumerate(r):
            draw.rectangle([x + c * col_w, y, x + (c + 1) * col_w - 2, y + row_h], fill=(22, 32, 52))
            draw.text((x + c * col_w + 8, y + 6), str(val)[:18], fill=MUTED, font=font(13))
        y += row_h + 1
    return y


def base(title: str, subtitle: str):
    img = Image.new("RGB", (W, H), BG)
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, W, 90], fill=PANEL)
    draw.text((36, 22), title, fill=TEXT, font=font(30, True))
    draw.text((36, 58), subtitle, fill=MUTED, font=font(16))
    draw.text((W - 280, 34), "Daniel Cohen Portfolio", fill=ACCENT, font=font(16, True))
    return img, draw


def kpi_card(draw, x, y, w, h, label, value, color=ACCENT):
    draw.rounded_rectangle([x, y, x + w, y + h], radius=14, fill=PANEL)
    draw.text((x + 18, y + 16), label, fill=MUTED, font=font(14))
    draw.text((x + 18, y + 44), str(value), fill=color, font=font(28, True))


def save(img: Image.Image, name: str):
    OUT.mkdir(parents=True, exist_ok=True)
    DEPLOY.mkdir(parents=True, exist_ok=True)
    for folder in (OUT, DEPLOY):
        path = folder / name
        img.save(path, "PNG", optimize=True)
        print(f"Wrote {path}")


def render_ai_dashboard():
    path = SQL / "AI_Evaluator_And_Analytics_Combined_Workbook.xlsx"
    rows = sheet_rows(path, "Dashboard", max_rows=10, max_cols=6)
    data = sheet_rows(path, "Evaluation_Data", max_rows=8, max_cols=7)
    img, draw = base(
        "AI Evaluator & Analytics Workbook",
        "KPI dashboard · prompt ratings · hallucination / safety metrics",
    )
    kpis = [
        ("Evaluations", "150", ACCENT),
        ("Avg Rating", "3.6", GREEN),
        ("Hallucination Flags", "18%", GOLD),
        ("SQL Queries", "11", ACCENT),
    ]
    for i, (label, value, color) in enumerate(kpis):
        kpi_card(draw, 36 + i * 380, 120, 360, 100, label, value, color)

    draw.rounded_rectangle([36, 250, W - 36, 520], radius=16, fill=PANEL)
    draw.text((56, 270), "Dashboard sheet (workbook KPIs)", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 310, rows if rows else [["Metric", "Value"], ["Total Evaluations", 150]], col_w=220)

    draw.rounded_rectangle([36, 540, W - 36, 960], radius=16, fill=PANEL)
    draw.text((56, 560), "Evaluation_Data sample", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 600, data, col_w=200)
    save(img, "sql-excel-ai-evaluator-dashboard.png")


def render_ai_sql():
    path = SQL / "AI_Evaluator_And_Analytics_Combined_Workbook.xlsx"
    rows = sheet_rows(path, "SQL_Practice", max_rows=14, max_cols=3)
    img, draw = base(
        "AI Evaluator — SQL Practice",
        "11 analytics queries · window functions · hallucination / model ranking",
    )
    draw.rounded_rectangle([36, 120, W - 36, 960], radius=16, fill=PANEL)
    draw.text((56, 140), "SQL_Practice sheet", fill=TEXT, font=font(18, True))
    draw.text(
        (56, 175),
        "Includes GROUP BY, CASE, HAVING, RANK() OVER, and subqueries on the evaluation dataset.",
        fill=MUTED,
        font=font(15),
    )
    draw_table(draw, 56, 220, rows if rows else [["#", "Query", "Focus"]], col_w=480, row_h=42)
    save(img, "sql-excel-ai-evaluator-sql.png")


def render_investment_dashboard():
    path = SQL / "Investment_App_Portfolio_Analysis_Workbook.xlsx"
    dash = sheet_rows(path, "Dashboard", max_rows=10, max_cols=6)
    risk = sheet_rows(path, "Risk_Metrics", max_rows=8, max_cols=6)
    img, draw = base(
        "Investment Portfolio Analysis Workbook",
        "Mirrors Streamlit app metrics · Sharpe · Monte Carlo · optimizer portfolios",
    )
    for i, (label, value, color) in enumerate(
        [
            ("Monte Carlo Sims", "500", ACCENT),
            ("Optimizer Candidates", "251", GREEN),
            ("Pivot Exercises", "10", GOLD),
            ("SQL Queries", "11", ACCENT),
        ]
    ):
        kpi_card(draw, 36 + i * 380, 120, 360, 100, label, value, color)

    draw.rounded_rectangle([36, 250, 780, 960], radius=16, fill=PANEL)
    draw.text((56, 270), "Dashboard", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 320, dash, col_w=160)

    draw.rounded_rectangle([810, 250, W - 36, 960], radius=16, fill=PANEL)
    draw.text((830, 270), "Risk_Metrics", fill=TEXT, font=font(18, True))
    draw_table(draw, 830, 320, risk, col_w=110)
    save(img, "sql-excel-investment-workbook-dashboard.png")


def render_investment_pivot():
    path = SQL / "Investment_App_Portfolio_Analysis_Workbook.xlsx"
    rows = sheet_rows(path, "Pivot_Practice", max_rows=14, max_cols=4)
    corr = sheet_rows(path, "Correlation_Data", max_rows=8, max_cols=7)
    img, draw = base(
        "Investment Workbook — Pivot Practice",
        "Allocation / risk pivot exercises · correlation matrix for portfolio analysis",
    )
    draw.rounded_rectangle([36, 120, 820, 960], radius=16, fill=PANEL)
    draw.text((56, 140), "Pivot_Practice", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 190, rows, col_w=180, row_h=36)

    draw.rounded_rectangle([850, 120, W - 36, 960], radius=16, fill=PANEL)
    draw.text((870, 140), "Correlation_Data", fill=TEXT, font=font(18, True))
    draw_table(draw, 870, 190, corr, col_w=90)
    save(img, "sql-excel-investment-pivot-practice.png")


def render_quant_claims():
    path = SQL / "Quant_Analytics_Portfolio_Workbook.xlsx"
    claims = sheet_rows(path, "Claims_Data", max_rows=12, max_cols=7)
    baseball = sheet_rows(path, "Baseball_Analytics", max_rows=8, max_cols=6)
    img, draw = base(
        "Quant Analytics Portfolio Workbook",
        "Insurance claims + baseball analytics · SQL and pivot practice",
    )
    kpi_card(draw, 36, 120, 360, 100, "Claims Records", "150", ACCENT)
    kpi_card(draw, 416, 120, 360, 100, "Baseball Records", "60", GREEN)
    kpi_card(draw, 796, 120, 360, 100, "SQL Questions", "10", GOLD)
    kpi_card(draw, 1176, 120, 388, 100, "Pivot Exercises", "8", ACCENT)

    draw.rounded_rectangle([36, 250, W - 36, 620], radius=16, fill=PANEL)
    draw.text((56, 270), "Claims_Data", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 310, claims, col_w=200)

    draw.rounded_rectangle([36, 650, W - 36, 960], radius=16, fill=PANEL)
    draw.text((56, 670), "Baseball_Analytics sample", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 710, baseball, col_w=180)
    save(img, "sql-excel-quant-claims.png")


def render_credit_risk():
    path = SQL / "Real_World_Data_Analytics_Portfolio_Workbook.xlsx"
    credit = sheet_rows(path, "Credit_Risk_Data", max_rows=12, max_cols=7)
    claims = sheet_rows(path, "Insurance_Claims", max_rows=8, max_cols=6)
    img, draw = base(
        "Real World Data Analytics Workbook",
        "Credit risk · insurance claims with fraud flags · portfolio returns",
    )
    kpi_card(draw, 36, 120, 360, 100, "Credit Companies", "200", ACCENT)
    kpi_card(draw, 416, 120, 360, 100, "Insurance Claims", "300", RED)
    kpi_card(draw, 796, 120, 360, 100, "SQL Tasks", "10", GOLD)
    kpi_card(draw, 1176, 120, 388, 100, "Domains", "3", GREEN)

    draw.rounded_rectangle([36, 250, W - 36, 620], radius=16, fill=PANEL)
    draw.text((56, 270), "Credit_Risk_Data", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 310, credit, col_w=200)

    draw.rounded_rectangle([36, 650, W - 36, 960], radius=16, fill=PANEL)
    draw.text((56, 670), "Insurance_Claims sample", fill=TEXT, font=font(18, True))
    draw_table(draw, 56, 710, claims, col_w=180)
    save(img, "sql-excel-real-world-credit-risk.png")


def main() -> int:
    render_ai_dashboard()
    render_ai_sql()
    render_investment_dashboard()
    render_investment_pivot()
    render_quant_claims()
    render_credit_risk()
    print("SQL/Excel screenshots complete")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
